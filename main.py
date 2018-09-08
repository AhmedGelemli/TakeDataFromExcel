import xlrd
from openpyxl import load_workbook

print("-----------------------------------------------------")
print()
print("Excel Data Claimer")
print("by Ahmed Gelemli")
print()
print("-----------------------------------------------------")

print('''
Options:
1)Take and see data in specific cell
2)Add sheet to the existing Excel files
3)Remove sheet from existing Excel file
4)Copy data from one sheet to another sheet
5)Write data (or Change)
''')

def opt_1():
    loc = (file_path) 
    wb = xlrd.open_workbook(loc) 
    sheet = wb.sheet_by_index(0) 
    row = int(input("Row:"))
    column = int(input("Column:"))
    value = sheet.cell_value(row, column)
    print(value)

def opt_2():
    filepath = file_path
    wb = load_workbook(filepath)
    sheet_name = input("Name of Sheet:")
    wb.create_sheet(sheet_name)
    wb.save(filepath)
    print("Your command is succesfully complated!")
    print(sheet_name + "has been added succesfully!")

def opt_3():
    filepath = file_path
    wb = load_workbook(filepath)
    sheet_rmv = input("Which sheet do you want to remove:")
    wb.remove(wb.get_sheet_by_name(sheet_rmv))
    wb.save(filepath)
    print("Your command is succesfully complated!")
    print(sheet_rmv + " has been removed!")

def opt_4():
    filepath = file_path
    wb = load_workbook(filepath)
    sheet_copy = input("Which sheet do you want to copy:")
    source = wb.get_sheet_by_name(sheet_copy)
    wb.copy_worksheet(source)
    wb.save(filepath)
    print()
    print("Your command is succesfully complated!")
    print("'" + sheet_copy + "' pasted as '" + sheet_copy + " Copy'!")

def opt_5():
    filepath = file_path
    wb=load_workbook(filepath)
    sheet=wb.active
    cell_num = input("Which cell do you want to change or write (ex: D3):")
    string = input("What do you want to write to " + cell_num + " :")
    sheet[cell_num] = string
    wb.save(filepath)
    print("Your command is succesfully complated!")
    print(string + " is writed to " + cell_num + "!")

 
option = input("Select Option:")
file_path = input("Excel File Path:")

if (option == "1"):
    opt_1()
elif (option == "2"):
    opt_2()
elif (option == "3"):
    opt_3()
elif (option == "4"):
    opt_4()
elif (option == "5"):
    opt_5()
else:
    print("Please select valid option!")